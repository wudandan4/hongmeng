# @Time     : 2022/3/1 18:20
# -*- coding: utf-8 -*-
import os
import random
import re
import time
from unicodedata import name
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook, Workbook
from selenium import webdriver
import openpyxl
import logging
import logging.handlers
import datetime
logger = logging.getLogger('mylogger')
logger.setLevel(logging.DEBUG)
rf_handler = logging.handlers.TimedRotatingFileHandler('yamaxun.log', when='midnight', interval=1, backupCount=7, atTime=datetime.time(0, 0, 0, 0))
rf_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(rf_handler)


def append_xlsx(path, sheetname, value):
    index = len(value)
    if not os.path.exists(path):
        wb = openpyxl.Workbook(path)
        wb.save(path)
        wb.close()
    time.sleep(1)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    for i in range(index):
        sheet.append(list(value[i]))
    workbook.save(path)
    print("xlsx格式表格数据追加成！")

class JingDong(object):
    def __init__(self, page_count):
        self.pc_agent = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 "
            "Safari/537.36",
            "Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) "
            "Version/5.1 Safari/534.50",
            "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 "
            "Safari/534.50",
            "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0);",
            "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
            "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
            "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
            "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
            "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 "
            "Safari/535.11",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; ",
            ".NET CLR 2.0.50727; SE 2.X MetaSr 1.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64; rv:76.0) Gecko/20100101 Firefox/76.0",
            'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1464.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/31.0.1650.16 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.3319.102 Safari/537.36',
            'Mozilla/5.0 (X11; CrOS i686 3912.101.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1667.0 Safari/537.36',
            'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:17.0) Gecko/20100101 Firefox/17.0.6',
            'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1468.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2224.3 Safari/537.36',
            'Mozilla/5.0 (X11; CrOS i686 3912.101.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36'
            ]
        self.all_url = [
            # 路由器
            'https://www.amazon.cn/b?ie=UTF8&node=1454010071'

        ]
        self.page_count = page_count

    def get_dianshiji_html_id(self):
        agent = random.choice(self.pc_agent)
        header = {'User-Agent': f"{agent}"}
        # 先去获取第一个url的信息
        # response = requests.get(self.all_url[0], headers=header).text
        # info = str(re.findall("_log:{wids:'(.*?)',uuid:", response)[0])
        # dianshiji_list = info.split(",")
        driver = webdriver.Chrome()
        agent = random.choice(self.pc_agent)
        href_list = []
        for i in range(1,10):
            link = 'https://www.amazon.com/s?k=TV&page='+str(i)
            driver.get(link)
            page = driver.page_source
            soup = BeautifulSoup(page, features='lxml')
            a_list = soup.find_all('a', {"class": "a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal"})
            a_set = set(a_list)
            for i in a_set:
                if '/dp/' in i.attrs['href']:
                    href = 'https://www.amazon.cn'+i.attrs['href']
                    href_list.append(href)

        print(href_list)
        return href_list


    

    def get_pages(self, list,chromedriver_path):
        
        end_info = set()
        options = webdriver.ChromeOptions()
        #处理ssl证书错误问题
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument('--log-level=1')
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)
        
        for url in list[36:]:
            logger.info("yamaxun3 url {}".format(url))
            
            # resultList = []
            # infoRes = []
            #link = 'https://www.amazon.cn/dp/ref=lp_1454010071_1_1'
            driver = webdriver.Chrome(executable_path = chromedriver_path, options = options)
            # driver.implicitly_wait(10) #隐式等待，网页加载数据需要时间，智能化等待
            agent = random.choice(self.pc_agent)
            header = {'User-Agent': f"{agent}"}
            driver.get(url)
            driver.implicitly_wait(10) #隐式等待，网页加载数据需要时间，智能化等待
            page = driver.page_source
            title = re.findall('<title>(.*?)</title>', page)
            # print("title {}".format(title))
            name = ''
            if len(title)>0:
                name = title[0]
                name = name.replace("亚马逊：商品评论:",'')
            
            soup = BeautifulSoup(page,features='lxml')
            try:
                review = soup.find('div',id='cm_cr-review_list')
                # print("review {}".format(review))
                review_text_content = review.findAll('span',class_='a-size-base review-text review-text-content')
                for content in review_text_content:
                    allList = []
                    infoList = []
                    infoList.append(name)
                    msg = content.get_text()
                    infoList.append(str(msg))
                    print("infoList {}".format(infoList))
                    allList.append(infoList)
                    pathfile = r'C:\project\luyouqi\result.xlsx'
                    append_xlsx(path=pathfile, sheetname='Sheet1', value=allList)
                    
                if '当前加载评论时遇到问题' in str(review):
                    with open("error.txt","a") as file:
                        print("url info {}".format(url))
                        file.write(url)
               
            except:
                pass

        # self.write_xlsx(name, title[0], end_info)
        

    def append_xlsx(path, sheetname, value):
        index = len(value)
        if not os.path.exists(path):
            wb = Workbook(path)
            wb.save(path)
            wb.close()
        time.sleep(1)
        workbook = load_workbook(path)
        sheet = workbook[sheetname]
        for i in range(index):
            sheet.append(list(value[i]))
        workbook.save(path)

if __name__ == '__main__':
    d = JingDong(2)
    # a = d.get_dianshiji_html_id()
    # d.bad_comment(a)
    # # d.write_xlsx('123','111', {1,2,3})
    chromedriver_path = r'C:\project\luyouqi\chromedriver_win32\chromedriver.exe'
    with open(r"C:\project\luyouqi\error.txt",'r') as f:
        URL_list = f.readlines()
    # detailList =['https://www.amazon.com/product-reviews/B085VM9ZDD/ref=cm_cr_dp_d_show_all_btm?ie=UTF8&reviewerType=all_reviews']
    # for detail in URL_list:
    #     detailList = []
    #     detailList.append(detail)
        # page2url = detail.replace('cm_cr_dp_d_show_all_btm','cm_cr_arp_d_paging_btm_next_2')
        # page2url = page2url.replace('all_reviews','all_reviews&pageNumber=2')
        # detailList.append(page2url)
        # page3url = detail.replace('cm_cr_dp_d_show_all_btm','cm_cr_getr_d_paging_btm_next_3')
        # page3url = page3url.replace('all_reviews','all_reviews&pageNumber=3')
        # detailList.append(page3url)
        # page4url = detail.replace('cm_cr_dp_d_show_all_btm','cm_cr_getr_d_paging_btm_next_4')
        # page4url = page4url.replace('all_reviews','all_reviews&pageNumber=4')
        # detailList.append(page4url)
        # d.get_pages(detailList,chromedriver_path)
    # detailList = ['https://www.amazon.com/Eero-6-Router/product-reviews/B085VM9ZDD/ref=cm_cr_getr_d_paging_btm_next_4?ie=UTF8&reviewerType=all_reviews&pageNumber=4']
    d.get_pages(URL_list,chromedriver_path)
  

    
