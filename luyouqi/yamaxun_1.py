# @Time     : 2022/3/1 18:20
# -*- coding: utf-8 -*-
import os
import random
import re
import time
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook, Workbook
from selenium import webdriver
import logging
import logging.handlers
import datetime
logger = logging.getLogger('mylogger')
logger.setLevel(logging.DEBUG)
rf_handler = logging.handlers.TimedRotatingFileHandler('yamaxun.log', when='midnight', interval=1, backupCount=7, atTime=datetime.time(0, 0, 0, 0))
rf_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(rf_handler)

class JingDong(object):
    def __init__(self, page_count):
        self.pc_agent = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 "
            "Safari/537.36",
            "Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) "
            "Version/5.1 Safari/534.50",
            "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 "
            "Safari/534.50",
            "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0);",
            "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
            "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
            "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
            "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
            "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 "
            "Safari/535.11",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; ",
            ".NET CLR 2.0.50727; SE 2.X MetaSr 1.0)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64; rv:76.0) Gecko/20100101 Firefox/76.0",
            'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1464.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/31.0.1650.16 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.3319.102 Safari/537.36',
            'Mozilla/5.0 (X11; CrOS i686 3912.101.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1667.0 Safari/537.36',
            'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:17.0) Gecko/20100101 Firefox/17.0.6',
            'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1468.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2224.3 Safari/537.36',
            'Mozilla/5.0 (X11; CrOS i686 3912.101.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36'
            ]
        self.all_url = [
            # 路由器
            'https://www.amazon.cn/b?ie=UTF8&node=1454010071'

        ]
        self.page_count = page_count


    def get_pages(self, list,chromedriver_path):
        
        end_info = set()
        options = webdriver.ChromeOptions()
        #处理ssl证书错误问题
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument('--log-level=1')
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)
        for url in list:
            logger.info("yamaxun_1 url : {}".format(url))
            if 'https' in url:
                resultList = []
                infoRes = []
                
                #link = 'https://www.amazon.cn/dp/ref=lp_1454010071_1_1'
                driver = webdriver.Chrome(executable_path = chromedriver_path, options = options)
                driver.implicitly_wait(10) #隐式等待，网页加载数据需要时间，智能化等待
                agent = random.choice(self.pc_agent)
                header = {'User-Agent': f"{agent}"}
                driver.get(url)
                driver.implicitly_wait(10) #隐式等待，网页加载数据需要时间，智能化等待
                page = driver.page_source
                soup = BeautifulSoup(page,features='lxml')
                try:
                    message_list = soup.find_all('a',class_='a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal')
                    # print("message_list",message_list)
                    for msg in message_list:
                        detail_url = msg.get("href")
                        print("detail_url {}".format(detail_url))
                        urlinfo = detail_url.replace("/-/zh/",'')
                        full_url = 'https://www.amazon.com/'+ urlinfo
                        print(full_url)
                        with open("origin.txt","a") as file:
                            file.write(full_url)
                            file.write('\n')
                except:
                    pass

        # self.write_xlsx(name, title[0], end_info)
        

    def append_xlsx(path, sheetname, value):
        index = len(value)
        if not os.path.exists(path):
            wb = Workbook(path)
            wb.save(path)
            wb.close()
        time.sleep(1)
        workbook = load_workbook(path)
        sheet = workbook[sheetname]
        for i in range(index):
            sheet.append(list(value[i]))
        workbook.save(path)

if __name__ == '__main__':
    d = JingDong(2)
    with open("C:\project\luyouqi\page.txt",'r') as f:
        URL_list = f.readlines()

    chromedriver_path = r'C:\project\luyouqi\chromedriver_win32\chromedriver.exe'
    # URL_list = ['https://www.amazon.com/s/?k=router&ref=sugsr_0_3&pd_rd_w=4CB2x&content-id=amzn1.sym.86e9859b-8a7b-4e73-aeaf-1df666c1aeb1:amzn1.sym.86e9859b-8a7b-4e73-aeaf-1df666c1aeb1&pf_rd_p=86e9859b-8a7b-4e73-aeaf-1df666c1aeb1&pf_rd_r=7AZSMSX00E5VY39N8QE8&pd_rd_wg=QnEWa&pd_rd_r=29566a0d-46db-4a37-baab-6d9f732bb460&qid=1663209846']
    # URL_list1 = ['https://www.amazon.com/TCL-4K-Smart-LED-43S435/product-reviews/B08DHDS4T3/ref=cm_cr_arp_d_viewopt_sr?ie=UTF8&reviewerType=all_reviews&filterByStar=one_star&pageSize=30&pageNumber=','https://www.amazon.com/TCL-40S325-Inch-1080p-Smart/product-reviews/B07GB61TQR/ref=cm_cr_getr_d_paging_btm_3?ie=UTF8&pageSize=30&filterByStar=one_star&reviewerType=all_reviews&pageNumber=','https://www.amazon.com/TCL-43-inch-Class-Smart-Android/product-reviews/B08P4YG1VQ/ref=cm_cr_arp_d_viewopt_sr?ie=UTF8&reviewerType=all_reviews&filterByStar=one_star&pageNumber=','https://www.amazon.com/TCL-32-inch-Class-Smart-Android/product-reviews/B08P4WR6XB/ref=cm_cr_arp_d_viewopt_sr?ie=UTF8&reviewerType=all_reviews&filterByStar=one_star&pageNumber=','https://www.amazon.com/TCL-Dolby-Vision-QLED-Smart/product-reviews/B08857ZHY4/ref=cm_cr_arp_d_paging_btm_next_2?ie=UTF8&reviewerType=all_reviews&pageNumber=','https://www.amazon.com/TCL-Dolby-Vision-QLED-Smart/product-reviews/B08C67986B/ref=cm_cr_getr_d_paging_btm_next_2?ie=UTF8&reviewerType=all_reviews&filterByStar=one_star&pageNumber=']
    
    d.get_pages(URL_list,chromedriver_path)
    # number = re.sub("\D","", 'a-icon a-icon-star a-star-5 review-rating')
    # print(type(number))

